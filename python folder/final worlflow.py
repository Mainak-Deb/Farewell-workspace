import openpyxl
import os
from PIL import Image

startportion='''
        <!DOCTYPE html>
        <html>
        <head>
            <meta http-equiv="Content-type" content="text/html;charset=UTF-8">
            <title>{name}</title>
            <link rel="stylesheet" href="//maxcdn.bootstrapcdn.com/bootstrap/3.2.0/css/bootstrap.min.css">
            <link rel="stylesheet" href="//maxcdn.bootstrapcdn.com/bootstrap/3.2.0/css/bootstrap-theme.min.css">
            <link rel="stylesheet" href="//code.jquery.com/ui/1.10.4/themes/smoothness/jquery-ui.css">
            <link rel='stylesheet' href="lib/isotope.css"/>
            <link rel='stylesheet' href="style.css"/>
            <style>

            </style>
        </head>
        <body>
        <div id="magazine">
        <div class="hard">
         <h1>Government College of Engineering & Textile Technolgy </h1>
          <h2>Serampore</h2>
      <h4>A Scrapbook<h4>

'''

endportion='''

          <div  class="hard"></div>
        <div  class="hard"></div>
      </div>
  </body>
  <script src="https://cdnjs.cloudflare.com/ajax/libs/jquery/3.3.1/jquery.min.js"></script>
  <script src="https://s3-us-west-2.amazonaws.com/s.cdpn.io/937765/turn.js"></script>


  <script>    
'use strict';

$('#magazine').turn({
  display:"double",
  width: 1200,
		height: 800,
  gradients: true, acceleration: true, turnCorners: "tl,tr"});


  </script>
</html>

'''

middleportion='''

middle

'''

print(startportion+ middleportion+endportion)


cover='''
<h4 style="font-size: 60px;" >{name}</h4>
<h4 style="font-size: 20px;" >{stream}</h4>
 </div>
  <div class="hard">
        </div>
         <div class="own-size scrapbook l">
          <div class="top-margin"></div>
          <p><b>The Ceremony</b></p>
          <p>On the 14th September, I was invited to Glaziers Hall on the South Bank to attend the 2018 London Regional Apprenticeship Awards, the building itself is right next to London Bridge.</p>
           <div class="glaziers-hall" style="  background-image: url('{singlepic}');"  ></div>
          <div class="london-bridge-postcard"  style="background-image: url('{grouppic}');"></div>
          <p></p>
          
        </div>
'''


leftpage='''
 <div class="own-size scrapbook l">
          <div class="top-margin"></div>
          <p style="text-align: center; font-family: 'Trirong', serif;" >
            <b>GCETTS page {pagenumber}</b> 
          </p><br><br>

        {maincontent}
          
</div>

'''

rightpage='''
 <div class="own-size scrapbook r">
          <div class="top-margin"></div>
          <p style="text-align: center; font-family: 'Trirong', serif;" >
            <b>GCETTS page {pagenumber}</b> 
          </p><br><br>

        {maincontent}
          
</div>
'''


paras='''
        <br><p><i>
        "{comment}"
        </i></p><br><br>
'''





c=paras.format(comment="this is working mainak")


print(leftpage.format(pagenumber=1,maincontent=c))



















def frmlrs(string):
  try:
    li=string.split()
    for i in range(len(li)):
        li[i]=li[i].capitalize()
    return "_".join(li)
  except:
    return ""



def dlrs(string):
  try:
    li=string.split()
    for i in range(len(li)):
        li[i]=li[i].capitalize()
    return " ".join(li)
  except:
    return ""


lineno=4



# path = "python folder\ITmemories.xlsx"
# save = 'C:/Users/hp/Documents/Farewell/IT scrapbook'



path = "python folder\CSEscrapbook.xlsx"
save = 'C:/Users/hp/Documents/Farewell/CSE htmls'



# path = "python folder\TT Scrapbook.xlsx"
# save = 'C:/Users/hp/Documents/Farewell/TT htmls'





wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

row=sheet_obj.max_row
column=sheet_obj.max_column

print(row,column)
li=[]
dep="Computer Science & Engineering"
for i in range(4,column):
    newname=sheet_obj.cell(row =1 , column = i).value
    newname=dlrs(newname);
    if(newname!=None):
        #print('/Images/'+str(newname)+'/'+str(newname)+'.jpg')
        names='Images\{}\{}.jpg'.format(str(newname),str(newname))
        print(names)
        #im =  open('Images\Ravi Ranjan Yadav\Ravi Ranjan Yadav.jpg', "r")
        try:
          im =  open(names, "r")
          cnt=cover.format(name=str(newname),stream=dep,singlepic='/Images/'+str(newname)+"/"+str(newname)+".jpg", grouppic='/Images/'+str(newname)+"/group-"+str(newname)+".jpg") 
        except:
          cnt=cover.format(name=str(newname),stream=dep,singlepic='https://bit.ly/3rICNF2', grouppic='https://images.static-collegedunia.com/public/college_data/images/appImage/1488542068c3.jpg?tr=c-force') 
        startportion2=startportion.format(name=str(newname)) 
        filename=str(frmlrs(newname))+".html"
        print(filename)        
        with open(os.path.join(save, filename), 'w', encoding='utf-8') as fp:
            a=[]
            pagecount=1;cmntcount=0;ss=''' ''';mains=''' '''
            for j in range(2,row):
                cell_obj = sheet_obj.cell(row = j, column = i)
                if(cell_obj.value!=None):
                  if((cmntcount!=0) and (cmntcount%lineno==0)):
                    pagecount+=1
                    if(pagecount%2==0):
                      mains=mains+leftpage.format(pagenumber=pagecount,maincontent=ss)
                      ss=''' '''
                      print("left",cmntcount,pagecount)
                    else:
                      mains=mains+rightpage.format(pagenumber=pagecount,maincontent=ss)
                      ss=''' '''
                      print("left",cmntcount,pagecount)
                  
                  cmntcount+=1;
                  ss=ss+paras.format(comment=str(cell_obj.value))

            if(pagecount%2==0):mains=mains+rightpage.format(pagenumber=pagecount+1,maincontent=ss)
            else:mains=mains+leftpage.format(pagenumber=pagecount+1,maincontent=ss)
                    #a.append(str(cell_obj.value)+"\n")
            fp.writelines(startportion2+cnt+mains+endportion )
            fp.close()
            pass
   
    li.append(a)



# Print value of cell object
# using the value attribute
print(li)
