import openpyxl

path = "python folder\CSEscrapbook.xlsx"

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

row=sheet_obj.max_row
column=sheet_obj.max_column

print(row,column)
li=[]

for i in range(1,column):
    a=[]
    for j in range(1,row):
        cell_obj = sheet_obj.cell(row = j, column = i)
        if(cell_obj.value!=None):
            a.append(cell_obj.value)
    li.append(a)



# Print value of cell object
# using the value attribute
print(li)
